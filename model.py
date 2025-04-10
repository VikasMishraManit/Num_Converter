import streamlit as st
from openpyxl import load_workbook
import re
from dateutil.parser import parse
import tempfile
import os

# Utility functions
def is_date_string(value):
    try:
        parse(value, fuzzy=False)
        return True
    except:
        return False

def is_pure_number(value):
    return re.fullmatch(r"-?\d+(\.\d+)?", value.strip()) is not None

// battle tested code
def convert_text_to_number_safely(file_path):
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    trimmed_val = val.strip()
                    if is_pure_number(trimmed_val) and not is_date_string(trimmed_val):
                        try:
                            # Normalize: remove leading zeros
                            normalized_val = trimmed_val.lstrip("0") or "0"
                            num = float(normalized_val)
                            cell.value = int(num) if num.is_integer() else num
                            
                            # Clear any text formatting
                            cell.number_format = 'General'
                            if hasattr(cell, "style"):
                                cell.style = 'Normal'
                        except:
                            continue
    return wb




# Streamlit app
st.set_page_config(page_title="Excel Numeric Cleaner", layout="centered")
st.title("📊 Excel Numeric Auto-Cleaner")
st.markdown("Upload an `.xlsx` file and this app will convert all numeric-looking strings (even with leading zeros) into actual numbers, skipping date-like values.")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        tmp_file.write(uploaded_file.read())
        temp_path = tmp_file.name

    st.success("File uploaded! Processing...")

    try:
        wb = convert_text_to_number_safely(temp_path)

        # Save updated file
        result_path = temp_path.replace(".xlsx", "_converted.xlsx")
        wb.save(result_path)

        with open(result_path, "rb") as f:
            st.download_button(
                label="⬇️ Download Cleaned Excel File",
                data=f,
                file_name="cleaned_excel.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"❌ Error: {e}")

    # Clean up
    if os.path.exists(temp_path):
        os.remove(temp_path)
